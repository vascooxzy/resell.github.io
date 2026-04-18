import io
import logging
import os
from pathlib import Path
from typing import Optional

import pandas as pd
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles

from scraper import scrape_dhgate
from sheets import export_to_sheets

# ─── Config ────────────────────────────────────────────────────────────────────

load_dotenv()

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s [%(levelname)s] %(name)s – %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)

# ─── App ───────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="DHgate Scraper API",
    description="Pesquisa e exporta produtos do DHgate.",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve frontend static files if the folder exists
frontend_dir = Path(__file__).parent.parent / "frontend"
if frontend_dir.exists():
    app.mount("/app", StaticFiles(directory=str(frontend_dir), html=True), name="frontend")


# ─── Routes ────────────────────────────────────────────────────────────────────

@app.get("/", tags=["health"])
def root():
    return {"status": "ok", "message": "DHgate Scraper API is running 🚀"}


@app.get("/search", tags=["scraping"])
def search(
    query: str = Query(..., min_length=1, description="Termo de pesquisa"),
    pages: int = Query(1, ge=1, le=5, description="Número de páginas a raspar"),
):
    """
    Pesquisa produtos no DHgate e devolve JSON.
    Guarda automaticamente um ficheiro Excel em /data/output.xlsx.
    """
    logger.info(f"Search request – query='{query}', pages={pages}")

    products = scrape_dhgate(query, max_pages=pages)

    if not products:
        raise HTTPException(
            status_code=404,
            detail="Nenhum produto encontrado. Tenta outro termo de pesquisa.",
        )

    # Auto-save Excel
    _save_excel(products, DATA_DIR / "output.xlsx")

    return {"query": query, "total": len(products), "products": products}


@app.get("/export/excel", tags=["export"])
def export_excel(
    query: str = Query(..., min_length=1),
    pages: int = Query(1, ge=1, le=5),
):
    """
    Raspa e devolve um ficheiro Excel para download directo.
    """
    logger.info(f"Excel export – query='{query}', pages={pages}")

    products = scrape_dhgate(query, max_pages=pages)

    if not products:
        raise HTTPException(status_code=404, detail="Nenhum produto encontrado.")

    output = _build_excel_bytes(products)
    filename = f"dhgate_{query.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/export/sheets", tags=["export"])
def export_sheets(
    query: str = Query(..., min_length=1),
    pages: int = Query(1, ge=1, le=5),
    sheet_id: Optional[str] = Query(None, description="ID do Google Spreadsheet"),
):
    """
    Raspa e envia os dados para o Google Sheets.
    """
    logger.info(f"Sheets export – query='{query}', sheet_id='{sheet_id}'")

    products = scrape_dhgate(query, max_pages=pages)

    if not products:
        raise HTTPException(status_code=404, detail="Nenhum produto encontrado.")

    try:
        url = export_to_sheets(products, sheet_id=sheet_id)
    except (FileNotFoundError, ValueError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("Google Sheets export failed")
        raise HTTPException(status_code=500, detail=f"Erro no Google Sheets: {e}")

    return {"url": url, "total": len(products)}


# ─── Helpers ───────────────────────────────────────────────────────────────────

def _build_excel_bytes(products: list[dict]) -> io.BytesIO:
    df = pd.DataFrame(products)
    df.columns = ["Nome", "Preço", "Avaliação", "Vendedor", "Link", "Imagem", "Encomendas"]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Produtos DHgate")
        ws = writer.sheets["Produtos DHgate"]

        # Column widths
        widths = {"A": 55, "B": 15, "C": 12, "D": 25, "E": 55, "F": 55, "G": 15}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # Header style
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", start_color="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

    output.seek(0)
    return output


def _save_excel(products: list[dict], path: Path) -> None:
    try:
        output = _build_excel_bytes(products)
        with open(path, "wb") as f:
            f.write(output.read())
        logger.info(f"Saved Excel to {path}")
    except Exception:
        logger.exception("Failed to save Excel file")
